import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

async def scrape_quartier_avec_clic(page, search_query):
    url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
    print(f"🔍 Scan approfondi (avec clics numéros) : {search_query}...")
    
    try:
        await page.goto(url)
        await page.wait_for_timeout(5000)
        
        # Validation des cookies Google
        try:
            button = await page.query_selector('button[aria-label="Tout accepter"]')
            if button:
                await button.click()
                await page.wait_for_timeout(2000)
        except Exception:
            pass

        # Scroll pour charger la liste de gauche
        panel_selector = 'div[role="feed"]'
        for _ in range(5): # On charge les premières agences du quartier
            await page.mouse.move(250, 400)
            await page.mouse.wheel(0, 3000)
            await page.wait_for_timeout(1000)
            
        elements = await page.query_selector_all('div[role="article"]')
        leads_quartier = []
        
        # On extrait les 15 premières agences de chaque quartier pour être ultra-précis sans blocage
        for index, el in enumerate(elements[:15]):
            try:
                title_el = await el.query_selector('div.fontHeadlineSmall')
                if not title_el:
                    continue
                name = await title_el.inner_text()
                
                # Clic obligatoire sur la fiche pour ouvrir le panneau de droite et charger les numéros cachés
                await el.click()
                await page.wait_for_timeout(2000) # Attente du chargement du panneau latéral droit
                
                # 1. Extraction du VRAI numéro de téléphone dans le panneau droit
                phone = "Non renseigné"
                phone_el = await page.query_selector('button[data-item-id^="phone:tel:"]')
                if phone_el:
                    phone_attr = await phone_el.get_attribute('data-item-id')
                    phone = phone_attr.replace("phone:tel:", "").strip()
                
                # 2. Extraction du VRAI Site Web
                has_website = "Non"
                website_el = await page.query_selector('a[data-item-id="authority"]')
                if website_el:
                    has_website = "Oui"

                # Notes et avis
                rating_el = await el.query_selector('span.MW4etd')
                rating = await rating_el.inner_text() if rating_el else "Aucune"
                
                reviews_el = await el.query_selector('span.UY7F9b')
                reviews_raw = await reviews_el.inner_text() if reviews_el else "0"
                reviews = int(reviews_raw.replace('(','').replace(')','').replace(' ','')) if reviews_raw != "0" else 0

                # Algorithme de priorité
                priority = "Basse"
                diagnostic = "Profil optimisé"
                
                if has_website == "Non":
                    priority = "Haute"
                    diagnostic = "Pas de site internet (Besoin urgent de Web Design)"
                elif rating != "Aucune" and float(rating.replace(",", ".")) < 4.0:
                    priority = "Haute"
                    diagnostic = f"Note critique ({rating}/5). Besoin d'E-Réputation"
                elif reviews < 10:
                    priority = "Moyenne"
                    diagnostic = "Manque de preuve sociale (Moins de 10 avis)"
                
                # Si on a le numéro, le lead est une opportunité en or
                leads_quartier.append({
                    "Nom de l'Agence": name,
                    "Quartier": search_query.replace("Agence immobilière ", ""),
                    "Téléphone": phone,
                    "Possède un Site Web": has_website,
                    "Note Google": rating,
                    "Nombre d'Avis": reviews,
                    "Priorité Démarchage": priority,
                    "Diagnostic Proposé": diagnostic
                })
                print(f"   ✅ Extrait : {name} | Tél: {phone} | Site: {has_website}")
                
            except Exception:
                continue
                
        return leads_quartier
    except Exception as e:
        print(f"⚠️ Erreur secteur : {e}")
        return []

def appliquer_style_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    high_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    high_font = Font(name="Calibri", size=11, bold=True, color="C00000")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            if col_idx == 6 and cell.value == "Haute":
                cell.fill = high_fill
                cell.font = high_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        
    wb.save(filename)

async def main():
    # On cible les secteurs majeurs de Rabat-salé pour avoir un maximum de leads qualifiés
    sectors = [
        "Agence immobilière Rabat Agdal", 
        "Agence immobilière Rabat Hay Riad", 
        "Agence immobilière Rabat Hassan", 
        "Agence immobilière Rabat Souissi", 
        "Agence immobilière Rabat Yacoub El Mansour", 
        "Agence immobilière Rabat Temara", 
        "Agence immobilière Rabat Oasis", 
        "Agence immobilière Rabat Centre Ville", 
        "Agence immobilière Rabat Hay Nahda",
        "Agence immobilière Rabat Hay Riad Temara",
        "Agence immobilière Rabat Hay Nahda",
        "Agence immobilière Salé Tabriquet ",
        "Agence immobilière Salé Bettana",
        "Agence immobilière Salé Maamora",
        "Agence immobilière Salé Kwass",
        "Agence immobilière Salé Hay Essalam",
        "Agence immobilière Salé Hay chrif",
        "Agence immobilière Salé Sidi Bouknadel",
        "Agence immobilière Salé Sidi Moussa",
        "Agence immobilière Salé Sidi Mohammed",
        "Agence immobilière Salé Hay Karima",
        "Agence immobilière Salé Hay Erhama",
        "Agence immobilière Salé Moulay Ismail",
        "Agence immobilière Bab Lamrissa",
        ]
    all_leads = []
    seen_names = set()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        
        for s in sectors:
            leads = await scrape_quartier_avec_clic(page, s)
            for lead in leads:
                if lead["Nom de l'Agence"] not in seen_names and not any(w in lead["Nom de l'Agence"].lower() for w in ["hotel", "guest"]):
                    seen_names.add(lead["Nom de l'Agence"])
                    all_leads.append(lead)
            print(f"📊 Total cumulé sans doublons : {len(all_leads)} agences.")
            print("-" * 40)
            
        await browser.close()
        
    if not all_leads:
        print("❌ Aucun lead trouvé.")
        return

    df = pd.DataFrame(all_leads)
    df['Sort_Order'] = df['Priorité Démarchage'].map({'Haute': 1, 'Moyenne': 2, 'Basse': 3})
    df = df.sort_values(by='Sort_Order').drop(columns=['Sort_Order'])
    
    filename = "Base_Prospects_Rabat_Premium.xlsx"
    df.to_excel(filename, index=False)
    appliquer_style_excel(filename)
    print(f"\n🏆 SCRAPING AVEC NUMÉROS TERMINE ! Fichier disponible : '{filename}'")

asyncio.run(main())