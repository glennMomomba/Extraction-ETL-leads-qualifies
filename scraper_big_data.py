import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Importation de la configuration centralisée
from config import Config

async def extraire_champ_secu(page, selector, attribute=None):
    """ Extrait une information en toute sécurité sans faire planter le script """
    try:
        element = await page.query_selector(selector)
        if element:
            if attribute:
                return await element.get_attribute(attribute)
            return await element.inner_text()
    except Exception:
        pass
    return "Non renseigné"

async def scrape_quartier_deep(page, search_query, nom_zone):
    url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
    print(f"🚀 Scan Deep Data pour : {nom_zone}...")
    
    try:
        await page.goto(url)
        await page.wait_for_timeout(Config.TIMEOUT)
        
        # Validation des cookies Google
        try:
            button = await page.query_selector('button[aria-label="Tout accepter"]')
            if button:
                await button.click()
                await page.wait_for_timeout(2000)
        except Exception:
            pass

        # Scroll la liste de gauche
        for _ in range(5):
            await page.mouse.move(250, 400)
            await page.mouse.wheel(0, 3000)
            await page.wait_for_timeout(1000)
            
        elements = await page.query_selector_all('div[role="article"]')
        leads_quartier = []
        
        for el in elements[:Config.MAX_LEADS]:
            try:
                title_el = await el.query_selector('div.fontHeadlineSmall')
                if not title_el:
                    continue
                name = await title_el.inner_text()
                
                link_el = await el.query_selector('a[href*="/maps/place/"]')
                maps_url = await link_el.get_attribute('href') if link_el else "Non renseigné"
                
                # Clic pour ouvrir le panneau latéral droit
                await el.click()
                await page.wait_for_timeout(2500)
                
                # 1. Extraction Téléphone
                phone = "Non renseigné"
                phone_el = await page.query_selector('button[data-item-id^="phone:tel:"]')
                if phone_el:
                    phone_attr = await phone_el.get_attribute('data-item-id')
                    phone = phone_attr.replace("phone:tel:", "").strip()
                
                # 2. Extraction Site Web
                site_url = "Non renseigné"
                web_el = await page.query_selector('a[data-item-id="authority"]')
                if web_el:
                    site_url = await web_el.get_attribute('href')
                
                # 3. Adresse physique
                adresse = await extraire_champ_secu(page, 'button[data-item-id^="address"]')
                adresse = adresse.replace("Adresse: ", "").strip() if adresse != "Non renseigné" else adresse
                
                # 4. Horaires
                horaires = await extraire_champ_secu(page, 'div[jsaction*="pane.info.hours"]')
                if horaires != "Non renseigné":
                    horaires = horaires.split("\n")[0]

                rating_el = await el.query_selector('span.MW4etd')
                rating = await rating_el.inner_text() if rating_el else "Aucune"
                
                reviews_el = await el.query_selector('span.UY7F9b')
                reviews_raw = await reviews_el.inner_text() if reviews_el else "0"
                reviews = int(reviews_raw.replace('(','').replace(')','').replace(' ','')) if reviews_raw != "0" else 0

                # Algorithme Business
                priority = "Basse"
                diagnostic = "Profil optimisé"
                
                if site_url == "Non renseigné":
                    priority = "Haute"
                    diagnostic = "Pas de site internet (Création de site vitrine urgente)"
                elif rating != "Aucune" and float(rating.replace(",", ".")) < 4.0:
                    priority = "Haute"
                    diagnostic = f"Note critique ({rating}/5). Besoin d'E-Réputation"
                elif reviews < 10:
                    priority = "Moyenne"
                    diagnostic = "Manque d'avis clients (Moins de 10 avis)"

                leads_quartier.append({
                    "Nom de l'Établissement": name,
                    "Zone / Requête": nom_zone,
                    "Téléphone": phone,
                    "Site Internet": site_url,
                    "Adresse Complète": adresse,
                    "Horaires / Statut": horaires,
                    "Note Google": rating,
                    "Nombre d'Avis": reviews,
                    "Priorité Démarchage": priority,
                    "Diagnostic Proposé": diagnostic,
                    "Lien Google Maps": maps_url
                })
                print(f"   ✅ Données profondes extraites pour : {name}")
                
            except Exception:
                continue
                
        return leads_quartier
    except Exception as e:
        print(f"⚠️ Erreur sur la zone {nom_zone} : {e}")
        return []

def appliquer_style_excel_premium(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    high_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    high_font = Font(name="Calibri", size=11, bold=True, color="C00000")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            
            if col_idx == 9 and cell.value == "Haute":
                cell.fill = high_fill
                cell.font = high_font
                
            if col_idx in [4, 11] and cell.value and "http" in str(cell.value):
                cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 40), 15)
        
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        
    wb.save(filename)

async def main():
    print(f"🚀 Démarrage du Data-Scraper — Thème actuel : '{Config.THEME}'")
    all_leads = []
    seen_names = set()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=Config.HEADLESS)
        page = await browser.new_page()
        await page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        
        # Parcourt la liste générée dynamiquement par config.py
        for s in Config.SECTORS:
            leads = await scrape_quartier_deep(page, s["query"], s["zone"])
            for lead in leads:
                if lead["Nom de l'Établissement"] not in seen_names:
                    seen_names.add(lead["Nom de l'Établissement"])
                    all_leads.append(lead)
            print(f"📊 Sous-total sans doublons : {len(all_leads)} fiches.")
            print("-" * 50)
            
        await browser.close()
        
    if not all_leads:
        print("❌ Aucun lead trouvé.")
        return

    df = pd.DataFrame(all_leads)
    df['Sort_Order'] = df['Priorité Démarchage'].map({'Haute': 1, 'Moyenne': 2, 'Basse': 3})
    df = df.sort_values(by=['Sort_Order', 'Zone / Requête']).drop(columns=['Sort_Order'])
    
    df.to_excel(Config.OUTPUT_FILENAME, index=False)
    appliquer_style_excel_premium(Config.OUTPUT_FILENAME)
    print(f"\n🏆 DATA ENRICHIE CHARGÉE ! Fichier Excel disponible : '{Config.OUTPUT_FILENAME}'")

if __name__ == "__main__":
    asyncio.run(main())